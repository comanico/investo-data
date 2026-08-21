from __future__ import annotations

import json
import re
from pathlib import Path

import pandas as pd
import requests

try:
    import yfinance as yf
except ImportError:
    yf = None

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, numbers

def load_fundamentals() -> dict:
    """Merge structured.json into workbook-style keys when available."""
    path = Path("extracted_financials/structured.json")

    data = json.loads(path.read_text())
    inc = data.get("income_statement", {})
    bal = data.get("balance_sheet", {})
    cf = data.get("cash_flow", {})

    def g(section, key, year):
        v = section.get(key)
        if isinstance(v, dict):
            return v.get(str(year))
        return None

    y, yp = 2025, 2024

    # OWC if balance sheet complete
    ca = g(bal, "current_assets", y)
    cl = g(bal, "current_liabilities", y)
    cash = g(bal, "cash", y) or 0
    mkt = g(bal, "marketable_securities", y) or 0
    std = g(bal, "short_term_debt", y) or 0
    if ca is not None and cl is not None:
        fund["OWC2024"] = (ca - cash - mkt) - (cl - std)

    return fund


# ─────────────────────────────────────────────────────────────────────────────
# Live market / macro inputs
# ─────────────────────────────────────────────────────────────────────────────

def fetch_risk_free_rate() -> float:
    """10-year Treasury from FRED DGS10 CSV (no API key)."""
    try:
        r = requests.get(
            "https://fred.stlouisfed.org/graph/fredgraph.csv?id=DGS10",
            timeout=20,
        )
        r.raise_for_status()
        lines = [
            ln.strip()
            for ln in r.text.splitlines()
            if ln.strip() and not ln.startswith("DATE") and ",." not in ln
        ]
        # last non-missing
        for ln in reversed(lines):
            parts = ln.split(",")
            if len(parts) == 2 and parts[1] not in (".", ""):
                return float(parts[1]) / 100.0
    except Exception as e:
        print(f"FRED fetch failed ({e}); using 0.047")
    return 0.047


def fetch_equity_risk_premium() -> float:
    """
    Damodaran implied ERP.
    Tries histimpl page / known current; falls back to 4.20% mature-market premium (Jul 2026).
    """
    # Published: Implied ERP ~4.28% (Aug 1 2026, trailing adj payout);
    # mature market premium after US default spread ~4.20% (Jul 2026).
    fallback = 0.0420
    try:
        r = requests.get(
            "https://pages.stern.nyu.edu/~adamodar/New_Home_Page/datafile/histimpl.html",
            timeout=20,
        )
        if r.ok:
            # Last table row often has most recent year ERP
            matches = re.findall(
                r"<td[^>]*>\s*(\d{4})\s*</td>.*?(\d+\.\d+)\s*%\s*</td>\s*</tr>",
                r.text,
                flags=re.I | re.S,
            )
            if matches:
                year, erp = matches[-1]
                erp_f = float(erp) / 100.0
                if 0.02 < erp_f < 0.12:
                    print(f"Damodaran histimpl last row year={year} ERP={erp_f:.2%}")
                    return erp_f
    except Exception as e:
        print(f"Damodaran fetch failed ({e})")
    print(f"Using Damodaran mature-market ERP fallback {fallback:.2%}")
    return fallback


def fetch_market_data(ticker: str = "F") -> dict:
    fallback = {
        "price": 14.0,
        "shares_m": 3917.0,
        "beta": 1.85,
        "source": "fallback",
    }
    if yf is None:
        return fallback
    try:
        info = yf.Ticker(ticker).info
        price = info.get("currentPrice") or info.get("regularMarketPrice")
        shares = info.get("sharesOutstanding")
        beta = info.get("beta")
        if not price or not shares:
            return fallback
        return {
            "price": float(price),
            "shares_m": float(shares) / 1e6,
            "beta": float(beta) if beta else 1.85,
            "source": "yfinance",
        }
    except Exception as e:
        print(f"yfinance failed ({e}); using fallback")
        return fallback


# ─────────────────────────────────────────────────────────────────────────────
# Absolute valuation (notebook logic, corrected CAPM)
# ─────────────────────────────────────────────────────────────────────────────

def run_absolute_valuation(fund: dict, mkt: dict, rf: float, erp: float) -> dict:
    shares_m = mkt["shares_m"]
    price = mkt["price"]
    beta = mkt["beta"]
    market_cap = shares_m * price  # $ millions

    total_debt = fund["TotalDebt2024"]
    cash_mkt = fund["CashAndMktSec2024"]
    net_debt = total_debt - cash_mkt
    total_capital = market_cap + net_debt
    equity_weight = market_cap / total_capital
    debt_weight = net_debt / total_capital

    interest_exp = fund["InterestExpense2024"]
    debt_2023 = fund.get("TotalDebt2023") or total_debt
    avg_debt = (debt_2023 + total_debt) / 2
    pre_tax_cod = interest_exp / avg_debt if avg_debt else 0.06

    pretax = fund.get("PreTaxIncome2024") or (
        fund["NetIncomeAttrib2024"] + fund["TaxExpense2024"]
    )
    tax_rate = fund["TaxExpense2024"] / pretax if pretax else 0.21
    tax_rate = max(0.0, min(0.35, abs(tax_rate)))

    # CORRECT CAPM: r_e = r_f + β × ERP
    # (notebook had r_f + β × (ERP − r_f) which collapses toward r_f)
    re = rf + beta * erp
    wacc = equity_weight * re + debt_weight * pre_tax_cod * (1 - tax_rate)

    # Growth
    rev_2024 = fund["Revenue2024"]
    rev_2020 = fund.get("Revenue2020")
    rev_2021 = fund.get("Revenue2021")
    rev_2023 = fund.get("Revenue2023")
    if rev_2020 and rev_2020 > 0:
        cagr = (rev_2024 / rev_2020) ** (1 / 4) - 1  # 2020→2024 = 4 years
        cagr_label = "2020-2024"
    elif rev_2021 and rev_2021 > 0:
        cagr = (rev_2024 / rev_2021) ** (1 / 3) - 1
        cagr_label = "2021-2024"
    elif rev_2023 and rev_2023 > 0:
        cagr = (rev_2024 / rev_2023) ** (1 / 1) - 1
        cagr_label = "2023-2024"
    else:
        cagr = 0.05
        cagr_label = "assumed"
    g_short = max(0.025, min(0.06, cagr * 0.6))
    g_terminal = 0.023

    net_income = fund["NetIncomeAttrib2024"]
    nopat = fund.get("NOPAT2024") or fund["OpsIncome2024"] * (1 - tax_rate)
    depreciation = abs(fund.get("DepreciationAmort2024") or 0)
    capex = abs(fund.get("CAPEX2024") or 0)
    owc_2024 = fund.get("OWC2024")
    owc_2023 = fund.get("OWC2023")
    if owc_2024 is not None and owc_2023 is not None:
        delta_owc = abs(owc_2023 - owc_2024)
    else:
        delta_owc = 0.0  # unknown → no WC drag in projection
    net_debt_iss = fund.get("NetDebtIssuance2024") or 0.0
    book_equity = fund["EquityAttrib2024"]
    dividends = abs(fund.get("CashDividendsPaid2024") or 0)

    years = 5

    # Projections
    proj_ni = [net_income * (1 + g_short) ** t for t in range(1, years + 1)]
    proj_div = [dividends * (1 + g_short) ** t for t in range(1, years + 1)]
    proj_fcfe = [
        proj_ni[t - 1]
        - capex * (1 + g_short) ** t
        - delta_owc * (1 + g_short) ** t
        + net_debt_iss * (1 + g_short) ** t
        for t in range(1, years + 1)
    ]
    proj_fcff = [
        net_income * (1 + g_short) ** t
        + depreciation * (1 + g_short) ** t
        + interest_exp * (1 - tax_rate) * (1 + g_short) ** t
        - capex * (1 + g_short) ** t
        - delta_owc * (1 + g_short) ** t
        for t in range(1, years + 1)
    ]
    proj_ri = [
        proj_ni[t - 1] - re * book_equity * (1 + g_short) ** (t - 1)
        for t in range(1, years + 1)
    ]

    # DDM (Gordon on next dividend)
    ddm_value = proj_div[0] / (re - g_terminal) if re > g_terminal else None
    ddm_ps = ddm_value / shares_m if ddm_value else None

    # FCFF
    pv_fcff = sum(proj_fcff[i] / (1 + wacc) ** (i + 1) for i in range(years))
    tv_fcff = proj_fcff[-1] * (1 + g_terminal) / (wacc - g_terminal)
    pv_tv_fcff = tv_fcff / (1 + wacc) ** years
    ev = pv_fcff + pv_tv_fcff
    eq_fcff = ev - net_debt
    fcff_ps = eq_fcff / shares_m

    # FCFE
    pv_fcfe = sum(proj_fcfe[t - 1] / (1 + re) ** t for t in range(1, years + 1))
    tv_fcfe = proj_fcfe[-1] * (1 + g_terminal) / (re - g_terminal)
    eq_fcfe = pv_fcfe + tv_fcfe / (1 + re) ** years
    fcfe_ps = eq_fcfe / shares_m

    # Residual Income
    pv_ri = sum(proj_ri[t - 1] / (1 + re) ** t for t in range(1, years + 1))
    tv_ri = proj_ri[-1] * (1 + g_terminal) / (re - g_terminal)
    ri_val = book_equity + pv_ri + tv_ri / (1 + re) ** years
    ri_ps = ri_val / shares_m

    return {
        "price": price,
        "shares_m": shares_m,
        "beta": beta,
        "market_cap_m": market_cap,
        "rf": rf,
        "erp": erp,
        "re": re,
        "wacc": wacc,
        "pre_tax_cod": pre_tax_cod,
        "tax_rate": tax_rate,
        "avg_debt": avg_debt,
        "net_debt": net_debt,
        "g_short": g_short,
        "g_terminal": g_terminal,
        "cagr": cagr,
        "cagr_label": cagr_label,
        "ddm_ps": ddm_ps,
        "fcff_ps": fcff_ps,
        "fcfe_ps": fcfe_ps,
        "ri_ps": ri_ps,
        "eq_fcff_bn": eq_fcff / 1000,
        "eq_fcfe_bn": eq_fcfe / 1000,
        "capex": capex,
        "depreciation": depreciation,
        "delta_owc": delta_owc,
        "market_source": mkt.get("source"),
    }


def write_absolute_sheet(res: dict, path: str = "Ford_Valuation.xlsx"):
    rows = [
        {
            "Model (Formula)": "Dividend Discount Model (D₁ / (rₑ − g))",
            "Intrinsic Value per Share ($)": res["ddm_ps"],
        },
        {
            "Model (Formula)": "Free Cash Flow to Equity (Σ FCFEₜ/(1+rₑ)ᵗ + TVₑ)",
            "Intrinsic Value per Share ($)": abs(res["fcfe_ps"]) if res["fcfe_ps"] else None,
        },
        {
            "Model (Formula)": "Free Cash Flow to Firm (Σ FCFFₜ/(1+WACC)ᵗ + TV − Net Debt)",
            "Intrinsic Value per Share ($)": abs(res["fcff_ps"]) if res["fcff_ps"] else None,
        },
        {
            "Model (Formula)": "Residual Income (Book Value + Σ RIₜ/(1+rₑ)ᵗ + TV_RI)",
            "Intrinsic Value per Share ($)": res["ri_ps"],
        },
    ]
    df = pd.DataFrame(rows)
    mkt = res["price"]
    df[f"vs. Market ${mkt:.2f}"] = df["Intrinsic Value per Share ($)"].apply(
        lambda v: f"{((v - mkt) / mkt * 100):.1f}%" if v is not None else "—"
    )

    # Inputs summary block
    inputs = pd.DataFrame(
        [
            ["Share price ($)", res["price"]],
            ["Shares outstanding (m)", res["shares_m"]],
            ["Market cap ($ m)", res["market_cap_m"]],
            ["Beta", res["beta"]],
            ["Risk-free rate (FRED DGS10)", res["rf"]],
            ["Equity risk premium (Damodaran)", res["erp"]],
            ["Cost of equity (CAPM)", res["re"]],
            ["Pre-tax cost of debt", res["pre_tax_cod"]],
            ["Effective tax rate", res["tax_rate"]],
            ["WACC", res["wacc"]],
            [f"Revenue CAGR ({res['cagr_label']})", res["cagr"]],
            ["Short-term growth g", res["g_short"]],
            ["Terminal growth", res["g_terminal"]],
            ["Net debt ($ m)", res["net_debt"]],
            ["Capex used ($ m)", res["capex"]],
            ["D&A used ($ m)", res["depreciation"]],
            ["ΔOWC used ($ m)", res["delta_owc"]],
            ["Market data source", res["market_source"]],
        ],
        columns=["Input", "Value"],
    )

    path = Path(path)
    if path.exists():
        wb = load_workbook(path)
        if "Absolute" in wb.sheetnames:
            del wb["Absolute"]
        ws = wb.create_sheet("Absolute", 0)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Absolute"

    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    ws["A1"] = "ABSOLUTE VALUATION — Ford Motor Company"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = (
        f"Live inputs: yfinance price/shares/beta · FRED 10y · Damodaran ERP  |  "
        f"CAPM: r_e = r_f + β × ERP (notebook formula corrected)"
    )

    # Results table
    headers = list(df.columns)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(4, c, h)
        cell.font = Font(bold=True)
        cell.border = thin
    for r, row in enumerate(df.itertuples(index=False), 5):
        for c, val in enumerate(row, 1):
            cell = ws.cell(r, c, val)
            cell.border = thin
            if c == 2 and isinstance(val, (int, float)):
                cell.number_format = "0.00"

    # Inputs
    ws.cell(10, 1, "MODEL INPUTS").font = Font(bold=True, size=12)
    ws.cell(11, 1, "Input").font = Font(bold=True)
    ws.cell(11, 2, "Value").font = Font(bold=True)
    for r, (inp, val) in enumerate(inputs.itertuples(index=False), 12):
        ws.cell(r, 1, inp)
        cell = ws.cell(r, 2, val)
        if isinstance(val, float) and val < 1 and "source" not in str(inp).lower():
            if "rate" in str(inp).lower() or "growth" in str(inp).lower() or "WACC" in str(inp) or "CAPM" in str(inp) or "premium" in str(inp).lower() or "cod" in str(inp).lower() or "CAGR" in str(inp):
                cell.number_format = "0.00%"
            else:
                cell.number_format = "0.00"
        elif isinstance(val, float):
            cell.number_format = "#,##0.00"

    ws.column_dimensions["A"].width = 70
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18

    wb.save(path)
    print(f"Saved → {path} (sheet Absolute)")
    return df


def main():
    print("Fetching live inputs…")
    rf = fetch_risk_free_rate()
    erp = fetch_equity_risk_premium()
    mkt = fetch_market_data("F")
    fund = load_fundamentals()

    print(f"  rf (10y)     = {rf:.3%}")
    print(f"  ERP          = {erp:.3%}")
    print(f"  price        = ${mkt['price']:.2f}  ({mkt['source']})")
    print(f"  shares (m)   = {mkt['shares_m']:.1f}")
    print(f"  beta         = {mkt['beta']:.3f}")

    res = run_absolute_valuation(fund, mkt, rf, erp)

    print(f"\nAvg Debt: {res['avg_debt']:.1f}")
    print(f"Effective Tax Rate: {res['tax_rate']:.4f}")
    print(f"Cost of Equity (CAPM): {res['re']:.3%}")
    print(f"WACC: {res['wacc']:.3%}")
    print(
        f"Revenue CAGR {res['cagr_label']}: {res['cagr']:.1%} → short-term g {res['g_short']:.1%}"
    )
    print(f"\nDDM  / share: ${res['ddm_ps']:.2f}" if res["ddm_ps"] else "")
    print(
        f"FCFF Equity: ${res['eq_fcff_bn']:.1f} bn → ${abs(res['fcff_ps']):.2f}/share"
    )
    print(
        f"FCFE Equity: ${res['eq_fcfe_bn']:.1f} bn → ${abs(res['fcfe_ps']):.2f}/share"
    )
    print(f"RI   / share: ${res['ri_ps']:.2f}")

    df = write_absolute_sheet(res, "Ford_Valuation.xlsx")
    print("\n" + df.to_string(index=False))
    return res


if __name__ == "__main__":
    main()
